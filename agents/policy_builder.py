import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import os
import json
import io
from langchain_groq import ChatGroq
from langchain_core.prompts import PromptTemplate
from dotenv import load_dotenv

load_dotenv()

GROQ_API_KEY = os.getenv("GROQ_API_KEY")
GROQ_MODEL   = os.getenv("GROQ_MODEL", "llama-3.3-70b-versatile")

POLICY_PROMPT = PromptTemplate(
    input_variables=[
        "company_name",
        "company_description",
        "information_flows",
        "key_stakeholders",
        "compliance_concerns",
        "existing_policies",
        "regulation_context",
    ],
    template="""
You are ARIRAS — a senior compliance advisor helping an Indian company understand
what they need to include in their compliance policy.

You do NOT assume any specific regulation. Instead, you look at:
1. The company's business profile and answers below
2. The regulation context extracted from their uploaded regulation document

Then you generate practical, plain-English guidance telling them WHAT to include,
WHY it matters, and a SAMPLE CLAUSE they can adapt.

COMPANY PROFILE:
- Company Name: {company_name}
- What the company does: {company_description}
- Information flows through the business: {information_flows}
- Key stakeholders: {key_stakeholders}
- Compliance areas of concern: {compliance_concerns}
- Existing policies / SOPs: {existing_policies}

REGULATION CONTEXT (extracted from uploaded regulation document):
{regulation_context}

TASK:
Based on the company profile AND the regulation context above:
1. Identify which regulation(s) apply to this company
2. Extract the key obligations from the regulation context
3. For each obligation, generate practical guidance for this specific company
4. Assign a readiness score (0-100) based on what they already have
5. Flag top priority actions and risk areas

Return ONLY valid JSON — no markdown, no explanation, no extra text:
{{
  "company_name": "{company_name}",
  "regulation_used": "<name of regulation identified from context, or General Best Practices if none>",
  "readiness_score": <integer 0-100>,
  "summary": "<2-3 sentence plain English summary of their compliance situation>",
  "sections": [
    {{
      "section_name": "<policy section name>",
      "what_to_include": "<plain English — exactly what this section must say for THIS company>",
      "sample_clause": "<a ready-to-adapt sample clause written for this company context>",
      "why_it_matters": "<simple explanation of the business or legal risk if this is missing>",
      "regulation_reference": "<exact clause or article from the regulation, e.g. Section 5, Article 13>"
    }}
  ],
  "priority_actions": [
    "<specific action this company must take immediately — be concrete, not generic>"
  ],
  "risk_areas": [
    "<specific risk area for this company based on their profile>"
  ]
}}

Generate 6-8 sections. Make every section specific to this company context —
not generic boilerplate. Reference the actual regulation clauses where possible.
""",
)


def generate_policy_guidance(
    company_name: str,
    company_description: str,
    information_flows: str,
    key_stakeholders: str,
    compliance_concerns: str,
    existing_policies: str,
    regulation_context: str = "",
) -> dict:

    llm = ChatGroq(
        model=GROQ_MODEL,
        api_key=GROQ_API_KEY,
        temperature=0.2,
    )

    prompt_text = POLICY_PROMPT.format(
        company_name=company_name,
        company_description=company_description,
        information_flows=information_flows,
        key_stakeholders=key_stakeholders,
        compliance_concerns=compliance_concerns,
        existing_policies=existing_policies,
        regulation_context=regulation_context or
            "No regulation uploaded. Use general compliance best practices.",
    )

    response = llm.invoke(prompt_text)
    raw = response.content.strip()

    if raw.startswith("```"):
        raw = raw.split("```")[1]
        if raw.startswith("json"):
            raw = raw[4:]
    raw = raw.strip()

    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        return {
            "company_name":    company_name,
            "regulation_used": "General",
            "readiness_score": 0,
            "summary":         "Could not parse AI response.",
            "sections": [{
                "section_name":         "Parse Error",
                "what_to_include":      raw[:300],
                "sample_clause":        "",
                "why_it_matters":       "Model returned unexpected format.",
                "regulation_reference": "",
            }],
            "priority_actions": [],
            "risk_areas":       [],
        }


def build_excel(guidance: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    BLUE_DARK   = "1E3A5F"
    BLUE_MID    = "2D6BE4"
    GREEN_DARK  = "0A4D3C"
    GREEN_LIGHT = "D1FAE5"
    AMBER_LIGHT = "FEF3C7"
    RED_LIGHT   = "FEE2E2"
    GREY_LIGHT  = "F8F9FA"
    WHITE       = "FFFFFF"
    TEXT_DARK   = "1A1D23"

    thin   = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hfont(size=10, color=WHITE):
        return Font(name="Arial", size=size, bold=True, color=color)

    def bfont(size=10, bold=False, color=TEXT_DARK):
        return Font(name="Arial", size=size, bold=bold, color=color)

    def fill(h):
        return PatternFill("solid", start_color=h, fgColor=h)

    def wrap(h="left"):
        return Alignment(wrap_text=True, vertical="top", horizontal=h)

    # ── Sheet 1: Policy Guidance ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Policy Guidance"

    ws1.merge_cells("A1:F1")
    ws1["A1"] = "ARIRAS — Compliance Policy Guidance Report"
    ws1["A1"].font      = Font(name="Arial", size=14, bold=True, color=WHITE)
    ws1["A1"].fill      = fill(BLUE_DARK)
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 32

    ws1.merge_cells("A2:F2")
    ws1["A2"] = (
        f"Company: {guidance.get('company_name','')}   |   "
        f"Regulation: {guidance.get('regulation_used','General')}   |   "
        f"Generated by ARIRAS"
    )
    ws1["A2"].font      = Font(name="Arial", size=10, color=WHITE)
    ws1["A2"].fill      = fill(BLUE_MID)
    ws1["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[2].height = 20

    ws1.merge_cells("A3:F3")
    ws1["A3"] = guidance.get("summary", "")
    ws1["A3"].font      = bfont(size=10, color="374151")
    ws1["A3"].fill      = fill("EFF6FF")
    ws1["A3"].alignment = wrap()
    ws1.row_dimensions[3].height = 45

    headers = [
        "Section Name", "What to Include", "Sample Clause",
        "Why It Matters", "Regulation Reference", "Priority",
    ]
    for col_idx, h in enumerate(headers, start=1):
        cell           = ws1.cell(row=4, column=col_idx)
        cell.value     = h
        cell.font      = hfont()
        cell.fill      = fill(BLUE_MID)
        cell.border    = border
        cell.alignment = wrap("center")
    ws1.row_dimensions[4].height = 22

    high_keywords = [
        "notice", "consent", "breach", "disclosure", "reporting",
        "penalty", "fine", "violation", "audit", "record", "register",
    ]

    sections = guidance.get("sections", [])
    for i, sec in enumerate(sections):
        combined = (
            sec.get("section_name", "") + " " +
            sec.get("what_to_include", "") + " " +
            sec.get("why_it_matters", "")
        ).lower()
        priority = "HIGH" if any(k in combined for k in high_keywords) else "MEDIUM"
        row_data = [
            sec.get("section_name", ""),
            sec.get("what_to_include", ""),
            sec.get("sample_clause", ""),
            sec.get("why_it_matters", ""),
            sec.get("regulation_reference", ""),
            priority,
        ]
        data_row = 5 + i
        ws1.append(row_data)
        row_bg = GREY_LIGHT if i % 2 == 0 else WHITE
        for col_idx in range(1, 7):
            cell           = ws1.cell(row=data_row, column=col_idx)
            cell.border    = border
            cell.alignment = wrap()
            if col_idx == 6:
                if priority == "HIGH":
                    cell.fill = fill(RED_LIGHT)
                    cell.font = bfont(bold=True, color="CC3333")
                else:
                    cell.fill = fill(AMBER_LIGHT)
                    cell.font = bfont(bold=True, color="CC8800")
            else:
                cell.fill = fill(row_bg)
                cell.font = bfont()
        ws1.row_dimensions[data_row].height = 80

    act_start = 5 + len(sections) + 2
    ws1.merge_cells(
        start_row=act_start, start_column=1,
        end_row=act_start,   end_column=6,
    )
    ws1.cell(row=act_start, column=1).value     = "⚡ Top Priority Actions"
    ws1.cell(row=act_start, column=1).font      = hfont(color=WHITE)
    ws1.cell(row=act_start, column=1).fill      = fill(GREEN_DARK)
    ws1.cell(row=act_start, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[act_start].height = 22

    for j, action in enumerate(guidance.get("priority_actions", []), start=1):
        r = act_start + j
        ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws1.cell(row=r, column=1).value     = f"{j}. {action}"
        ws1.cell(row=r, column=1).font      = bfont()
        ws1.cell(row=r, column=1).fill      = fill(GREEN_LIGHT)
        ws1.cell(row=r, column=1).alignment = wrap()
        ws1.row_dimensions[r].height        = 28

    for i, w in enumerate([28, 45, 55, 40, 28, 12], start=1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 2: Summary & Score ──────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary & Score")

    ws2.merge_cells("A1:D1")
    ws2["A1"] = "ARIRAS — Compliance Readiness Summary"
    ws2["A1"].font      = Font(name="Arial", size=13, bold=True, color=WHITE)
    ws2["A1"].fill      = fill(BLUE_DARK)
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 30

    score    = guidance.get("readiness_score", 0)
    sc_color = "0A4D3C" if score >= 70 else "8B5000" if score >= 40 else "7F1D1D"
    sc_bg    = GREEN_LIGHT if score >= 70 else AMBER_LIGHT if score >= 40 else RED_LIGHT

    ws2.merge_cells("A2:D2")
    ws2["A2"] = f"Compliance Readiness Score:  {score} / 100"
    ws2["A2"].font      = Font(name="Arial", size=16, bold=True, color=sc_color)
    ws2["A2"].fill      = fill(sc_bg)
    ws2["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[2].height = 36

    ws2.merge_cells("A3:D3")
    ws2["A3"] = guidance.get("summary", "")
    ws2["A3"].font      = bfont(size=10, color="374151")
    ws2["A3"].fill      = fill("EFF6FF")
    ws2["A3"].alignment = wrap()
    ws2.row_dimensions[3].height = 50

    info_rows = [
        ("Company Name",     guidance.get("company_name", "")),
        ("Regulation Used",  guidance.get("regulation_used", "")),
        ("Sections Covered", str(len(sections))),
        ("Priority Actions", str(len(guidance.get("priority_actions", [])))),
        ("Risk Areas",       str(len(guidance.get("risk_areas", [])))),
    ]
    for c, label in [(1, "Profile"), (2, "Details")]:
        cell = ws2.cell(row=4, column=c)
        cell.value, cell.font, cell.fill = label, hfont(), fill(BLUE_MID)
        cell.border, cell.alignment = border, wrap("center")
    ws2.row_dimensions[4].height = 20

    for idx, (label, value) in enumerate(info_rows, start=5):
        for c, v in [(1, label), (2, value)]:
            cell           = ws2.cell(row=idx, column=c)
            cell.value     = v
            cell.font      = bfont()
            cell.fill      = fill(GREY_LIGHT if idx % 2 == 0 else WHITE)
            cell.border    = border
            cell.alignment = wrap()
        ws2.row_dimensions[idx].height = 20

    risk_start = 5 + len(info_rows) + 2
    ws2.merge_cells(
        start_row=risk_start, start_column=1,
        end_row=risk_start,   end_column=4,
    )
    ws2.cell(row=risk_start, column=1).value     = "⚠️  Key Risk Areas"
    ws2.cell(row=risk_start, column=1).font      = hfont(color=WHITE)
    ws2.cell(row=risk_start, column=1).fill      = fill("7F1D1D")
    ws2.cell(row=risk_start, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[risk_start].height = 22

    for k, risk in enumerate(guidance.get("risk_areas", []), start=1):
        r = risk_start + k
        ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        ws2.cell(row=r, column=1).value     = f"• {risk}"
        ws2.cell(row=r, column=1).font      = bfont()
        ws2.cell(row=r, column=1).fill      = fill(RED_LIGHT)
        ws2.cell(row=r, column=1).alignment = wrap()
        ws2.row_dimensions[r].height        = 28

    for col, w in [("A", 30), ("B", 50), ("C", 20), ("D", 20)]:
        ws2.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()