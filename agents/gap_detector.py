import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import os
import json
import io
from langchain_groq import ChatGroq
from langchain_core.prompts import PromptTemplate
from langchain_community.document_loaders import PyPDFLoader
from core.vectorstore import load_vectorstore
from core.edge_handler import preflight_gap, gap_fallback_response  # ← NEW
from dotenv import load_dotenv

load_dotenv()

GROQ_API_KEY = os.getenv("GROQ_API_KEY")
GROQ_MODEL   = os.getenv("GROQ_MODEL", "llama-3.3-70b-versatile")

GAP_PROMPT = PromptTemplate(
    input_variables=["regulation_clauses", "policy_text"],
    template="""
You are ARIRAS — an expert AI compliance auditor.

Your job is to compare a COMPANY POLICY against REGULATION CLAUSES and produce a detailed gap analysis report.

REGULATION CLAUSES (extracted from uploaded regulation):
{regulation_clauses}

COMPANY POLICY TEXT:
{policy_text}

INSTRUCTIONS:
1. Extract every key obligation from the regulation clauses.
2. Check each obligation against the company policy — determine: MET, PARTIAL, or GAP.
3. For each GAP or PARTIAL, provide:
   - A short obligation name
   - Severity: high, medium, or low
   - What the regulation actually says (the source clause — quote or paraphrase the exact requirement)
   - What is missing or wrong in the company policy
   - A clear rationale explaining WHY this is a gap and what the business/legal risk is if not fixed
4. For obligations that are MET, just list the name.
5. Calculate a compliance score (0-100).

Return ONLY valid JSON — no markdown, no explanation:
{{
  "compliance_score": <integer 0-100>,
  "regulation_name": "<name of regulation from context>",
  "gaps": [
    {{
      "obligation": "<short obligation name>",
      "severity": "<high|medium|low>",
      "regulation_source": "<what the regulation actually requires>",
      "what_is_missing": "<what is absent or incorrect in the company policy>",
      "rationale": "<plain English explanation of why this matters>",
      "recommended_action": "<one concrete fix>"
    }}
  ],
  "met": [
    {{
      "obligation": "<obligation name>",
      "note": "<brief note on how the policy satisfies this>"
    }}
  ]
}}
""",
)


def detect_gaps(policy_file=None, policy_text: str = None, regulation_name: str = "") -> dict:
    llm = ChatGroq(model=GROQ_MODEL, api_key=GROQ_API_KEY, temperature=0)

    if policy_text is None and policy_file is not None:
        os.makedirs("data/uploads", exist_ok=True)
        policy_path = f"data/uploads/policy_{policy_file.name}"
        with open(policy_path, "wb") as f:
            f.write(policy_file.read())
        if policy_file.name.endswith(".pdf"):
            loader      = PyPDFLoader(policy_path)
            pages       = loader.load()
            policy_text = "\n".join([p.page_content for p in pages])[:4000]
        else:
            with open(policy_path, "r", encoding="utf-8", errors="ignore") as f:
                policy_text = f.read()[:4000]

    vectorstore = load_vectorstore()
    retriever   = vectorstore.as_retriever(search_kwargs={"k": 6})
    reg_docs    = retriever.invoke("key obligations reporting requirements disclosure rules")
    reg_clauses = "\n\n".join([d.page_content for d in reg_docs])[:3000]

    # ── Semantic preflight — runs before spending an LLM call ────────────────
    preflight = preflight_gap(policy_text or "", reg_docs)
    if not preflight["ok"]:
        fallback             = gap_fallback_response(
            reason           = "; ".join(preflight["errors"]),
            regulation_name  = regulation_name,
        )
        fallback["_preflight"] = preflight
        return fallback

    prompt_text = GAP_PROMPT.format(
        regulation_clauses=reg_clauses,
        policy_text=policy_text or "No policy text provided.",
    )

    response = llm.invoke(prompt_text)
    raw      = response.content.strip()

    if raw.startswith("```"):
        raw = raw.split("```")[1]
        if raw.startswith("json"):
            raw = raw[4:]
    raw = raw.strip()

    try:
        result = json.loads(raw)
    except json.JSONDecodeError:
        result = {
            "compliance_score": 0,
            "regulation_name":  regulation_name,
            "gaps": [{
                "obligation":         "Parsing error",
                "severity":           "high",
                "regulation_source":  "",
                "what_is_missing":    f"Model returned non-JSON: {raw[:200]}",
                "rationale":          "Could not parse model output.",
                "recommended_action": "Re-run the analysis.",
            }],
            "met": [],
        }

    result["_preflight"] = preflight  # attach for UI warnings
    return result


# ─────────────────────────────────────────────────────────────────────────────
# build_gap_excel — completely unchanged from original
# ─────────────────────────────────────────────────────────────────────────────
def build_gap_excel(result: dict, regulation_name: str = "", policy_name: str = "") -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    BLUE_DARK   = "1E3A5F"; BLUE_MID    = "2D6BE4"
    RED_LIGHT   = "FEE2E2"; AMBER_LIGHT = "FEF3C7"
    GREEN_DARK  = "064E3B"; GREEN_LIGHT = "D1FAE5"
    GREY_LIGHT  = "F8F9FA"; WHITE       = "FFFFFF"; TEXT_DARK = "1A1D23"
    thin   = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    def hfont(size=10, color=WHITE): return Font(name="Arial", size=size, bold=True, color=color)
    def bfont(size=10, bold=False, color=TEXT_DARK): return Font(name="Arial", size=size, bold=bold, color=color)
    def fill(h): return PatternFill("solid", start_color=h, fgColor=h)
    def wrap(h="left"): return Alignment(wrap_text=True, vertical="top", horizontal=h)

    score    = result.get("compliance_score", 0)
    gaps     = result.get("gaps", [])
    met      = result.get("met", [])
    reg_used = result.get("regulation_name", regulation_name or "General")
    sc_color = "0A4D3C" if score >= 70 else "8B5000" if score >= 40 else "7F1D1D"
    sc_bg    = GREEN_LIGHT if score >= 70 else AMBER_LIGHT if score >= 40 else RED_LIGHT
    import datetime; generated = datetime.datetime.now().strftime("%d %b %Y, %H:%M")

    ws1 = wb.active; ws1.title = "Gap Analysis"
    ws1.merge_cells("A1:G1"); ws1["A1"] = "ARIRAS — Compliance Gap Analysis Report"
    ws1["A1"].font = Font(name="Arial", size=14, bold=True, color=WHITE)
    ws1["A1"].fill = fill(BLUE_DARK); ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 32
    ws1.merge_cells("A2:G2")
    ws1["A2"] = f"Policy: {policy_name or 'Uploaded Document'}   |   Regulation: {reg_used}   |   Generated: {generated}   |   ARIRAS"
    ws1["A2"].font = Font(name="Arial", size=10, color=WHITE); ws1["A2"].fill = fill(BLUE_MID)
    ws1["A2"].alignment = Alignment(horizontal="center", vertical="center"); ws1.row_dimensions[2].height = 18
    ws1.merge_cells("A3:G3")
    ws1["A3"] = f"Compliance Score: {score}/100   |   Gaps: {len(gaps)}   |   Obligations Met: {len(met)}"
    ws1["A3"].font = Font(name="Arial", size=12, bold=True, color=sc_color)
    ws1["A3"].fill = fill(sc_bg); ws1["A3"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[3].height = 26; ws1.append([])
    headers = ["#","Obligation","Severity","Regulation Source / Clause","What Is Missing","Rationale","Recommended Action"]
    for ci, h in enumerate(headers, 1):
        c = ws1.cell(row=5, column=ci); c.value = h; c.font = hfont(); c.fill = fill(BLUE_MID)
        c.border = border; c.alignment = wrap("center")
    ws1.row_dimensions[5].height = 24
    sev_fill = {"high": RED_LIGHT, "medium": AMBER_LIGHT, "low": GREEN_LIGHT}
    sev_font = {"high": "CC3333",  "medium": "CC8800",    "low": "065F46"}
    for i, gap in enumerate(gaps):
        sev = gap.get("severity", "medium").lower()
        ws1.append([i+1, gap.get("obligation",""), gap.get("severity","medium").upper(),
                    gap.get("regulation_source",""), gap.get("what_is_missing",""),
                    gap.get("rationale",""), gap.get("recommended_action","")])
        dr = 6 + i; row_bg = GREY_LIGHT if i % 2 == 0 else WHITE
        for ci in range(1, 8):
            c = ws1.cell(row=dr, column=ci); c.border = border; c.alignment = wrap()
            if ci == 1: c.font = bfont(bold=True); c.fill = fill(row_bg); c.alignment = wrap("center")
            elif ci == 3: c.fill = fill(sev_fill.get(sev, AMBER_LIGHT)); c.font = Font(name="Arial", size=10, bold=True, color=sev_font.get(sev,"CC8800")); c.alignment = wrap("center")
            else: c.fill = fill(row_bg); c.font = bfont()
        ws1.row_dimensions[dr].height = 80
    for i, w in enumerate([5,28,12,40,42,48,40], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    ws2 = wb.create_sheet("Obligations Met")
    ws2.merge_cells("A1:C1"); ws2["A1"] = "ARIRAS — Obligations Already Met"
    ws2["A1"].font = Font(name="Arial", size=13, bold=True, color=WHITE); ws2["A1"].fill = fill(GREEN_DARK)
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center"); ws2.row_dimensions[1].height = 28
    ws2.merge_cells("A2:C2"); ws2["A2"] = f"Policy: {policy_name or 'Uploaded Document'}   |   Regulation: {reg_used}"
    ws2["A2"].font = Font(name="Arial", size=10, color=WHITE); ws2["A2"].fill = fill(BLUE_MID)
    ws2["A2"].alignment = Alignment(horizontal="center", vertical="center"); ws2.row_dimensions[2].height = 18
    ws2.append([])
    for c, h in [(1,"#"),(2,"Obligation"),(3,"How Your Policy Satisfies This")]:
        cell = ws2.cell(row=4, column=c); cell.value = h; cell.font = hfont(); cell.fill = fill(GREEN_DARK)
        cell.border = border; cell.alignment = wrap("center")
    ws2.row_dimensions[4].height = 22
    for idx, item in enumerate(met):
        obl, note = (item.get("obligation",""), item.get("note","Satisfies the requirement")) if isinstance(item, dict) else (str(item), "Satisfies the requirement")
        r = 5 + idx; bg = GREY_LIGHT if idx % 2 == 0 else WHITE
        for c, v in [(1, idx+1),(2, obl),(3, note)]:
            cell = ws2.cell(row=r, column=c); cell.value = v; cell.font = bfont(bold=(c==1))
            cell.fill = fill(bg); cell.border = border; cell.alignment = wrap("center" if c==1 else "left")
        ws2.row_dimensions[r].height = 40
    ws2.column_dimensions["A"].width = 5; ws2.column_dimensions["B"].width = 40; ws2.column_dimensions["C"].width = 60

    ws3 = wb.create_sheet("Summary")
    ws3.merge_cells("A1:B1"); ws3["A1"] = "ARIRAS — Compliance Summary"
    ws3["A1"].font = Font(name="Arial", size=13, bold=True, color=WHITE); ws3["A1"].fill = fill(BLUE_DARK)
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center"); ws3.row_dimensions[1].height = 28
    ws3.merge_cells("A2:B2"); ws3["A2"] = f"Compliance Score: {score} / 100"
    ws3["A2"].font = Font(name="Arial", size=16, bold=True, color=sc_color); ws3["A2"].fill = fill(sc_bg)
    ws3["A2"].alignment = Alignment(horizontal="center", vertical="center"); ws3.row_dimensions[2].height = 34
    summary_rows = [("Policy Analysed", policy_name or "Uploaded Document"),("Regulation Used", reg_used),
        ("Total Gaps", str(len(gaps))),("High Severity", str(sum(1 for g in gaps if g.get("severity")=="high"))),
        ("Medium Severity", str(sum(1 for g in gaps if g.get("severity")=="medium"))),
        ("Low Severity", str(sum(1 for g in gaps if g.get("severity")=="low"))),
        ("Obligations Met", str(len(met))),("Generated At", generated)]
    for c, h in [(1,"Field"),(2,"Value")]:
        cell = ws3.cell(row=3, column=c); cell.value = h; cell.font = hfont(); cell.fill = fill(BLUE_MID)
        cell.border = border; cell.alignment = wrap("center")
    ws3.row_dimensions[3].height = 20
    for idx, (label, value) in enumerate(summary_rows, 4):
        bg = GREY_LIGHT if idx % 2 == 0 else WHITE
        for c, v in [(1,label),(2,value)]:
            cell = ws3.cell(row=idx, column=c); cell.value = v; cell.font = bfont(bold=(c==1))
            cell.fill = fill(bg); cell.border = border; cell.alignment = wrap()
        ws3.row_dimensions[idx].height = 22
    ws3.column_dimensions["A"].width = 30; ws3.column_dimensions["B"].width = 45

    buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf.getvalue()